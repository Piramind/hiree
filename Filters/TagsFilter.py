# -*- coding: utf-8 -*-

from .ParentFilter import *
import os
from openpyxl import load_workbook


class TagsFilter(ParentFilter):
    __metaclass__ = ABCMeta

    def __init__(self, file_name: str, wb_name: str, main: bool, about_myself: bool):
        super().__init__(file_name)

        wb_path = os.getcwd() + '/' + wb_name
        wb = load_workbook(filename=wb_path, read_only=True)
        ws = wb["Лист2"]  # временный костыль

        self.main, self.about_myself = main, about_myself

        self.keywords1 = set(self._get_keywords(ws['A2':'A100']))
        self.keywords2 = set(self._get_keywords(ws['B2':'B100']))
        self.keywords3 = set(self._get_keywords(ws['C2':'C100']))
        self.keywords4 = set(self._get_keywords(ws['D2':'D100']))
        self.keywords5 = set(self._get_keywords(ws['E2':'E100']))

    def _get_keywords(self, cell_range: tuple) -> tuple:
        keywords = ()
        for row in cell_range:
            for cell in row:
                cell_value = cell.value
                if not cell_value:
                    break
                keywords += (cell_value,)
        return keywords

    # Запуск фильтра
    @abstractmethod
    def run(self) -> None:
        super().run()
