# -*- coding: utf-8 -*-
from Filters.MakeReport import MakeReport
from tqdm import tqdm
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class hhMakeReport(MakeReport):
    def __init__(self, file_name: str = "hh_RESULT.txt", wb_name: str = "hh_REPORT.xlsx"):
        super().__init__(file_name, wb_name)

    def run(self) -> None:
        print("HeadHunter: Формируем отчёт...")

        wb_path = os.getcwd() + '/' + self.wb_name
        wb = Workbook()
        wb.save(wb_path)
        wb = load_workbook(filename=wb_path)

        with open(self.file_name, 'r', encoding='utf-8') as file:
            progress = int(file.readline().strip())
            ws = wb.active

            for resume_ind in range(progress):
                column_letter = get_column_letter(1+resume_ind*3)
                ws[column_letter+'1'] = 'ФИО:'
                ws[column_letter+'2'] = 'Резюме:'

            for resume_ind in range(progress):
                column_letter = get_column_letter(2+resume_ind*3)

                link = file.readline().strip()
                ws[column_letter+'2'] = link

                # html = super()._get_html(link)
                # soup = BeautifulSoup(html, 'lxml')

        wb.save(wb_path)
